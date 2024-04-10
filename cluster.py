
# Impotation des bibliothèques


import cv2
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.cluster import DBSCAN
import tkinter as tk
from tkinter import messagebox
import openpyxl
from openpyxl.styles import PatternFill




def cluster_fishes_DBSCAN(fish_centers, eps, min_samples):
    """
    Effectue le clustering des poissons en utilisant l'algorithme DBSCAN.

    @input :
    - fish_centers (dict): Dictionnaire contenant les coordonnées des poissons pour chaque trame.
    - eps (float): Rayon de la sphère autour d'un point pour rechercher les voisins.
    - min_samples (int): Nombre minimum de points requis pour former un cluster.

    @output :
    - all_frames_clusters (dict): Dictionnaire contenant les informations sur les clusters pour chaque trame.
    """
    all_frames_clusters = {}
    for frame_number, frame_data in fish_centers.items():
        fish_coordinates = [frame_data[str(fish_id)][0] for fish_id in frame_data]

        # Utilisation de DBSCAN
        dbscan = DBSCAN(eps=eps, min_samples=min_samples)
        cluster_labels = dbscan.fit_predict(fish_coordinates)

        # Création des clusters
        unique_labels = set(cluster_labels)
        num_clusters = len(unique_labels) - (1 if -1 in unique_labels else 0)  # Ignorer le label -1 pour le bruit
        fish_clusters = {str(i + 1): [] for i in range(num_clusters)}
        for i, label in enumerate(cluster_labels):
            if label != -1:  # Ignorer le bruit
                fish_clusters[str(label + 1)].append((str(i + 1), fish_coordinates[i]))

        # Construction des informations sur les clusters
        cluster_info = {}
        for cluster, fishes in fish_clusters.items():
            cluster_info[cluster] = {
                "Fish Positions": [fish[1] for fish in fishes],
                "Cluster Center": [sum(pos) / len(pos) for pos in zip(*[fish[1] for fish in fishes])]
            }

        all_frames_clusters[frame_number] = {"Clusters": cluster_info}

    return all_frames_clusters


def draw_clusters_on_image(image, cluster_data):
    """
    Dessine les clusters sur une seule image.

    @input :
    - image (numpy.ndarray): Image sur laquelle dessiner les clusters.
    - cluster_data (dict): Informations sur les clusters à dessiner.

    @output :
    - frame_with_clusters (numpy.ndarray): Image avec les clusters dessinés.
    """
    frame_with_clusters = image.copy()  # Copie de l'image d'origine pour éviter de la modifier directement
    for cluster_info in cluster_data.values():  # Parcours de chaque cluster dans les données
        cluster_center = tuple(map(int, cluster_info["Cluster Center"]))  # Centre du cluster (converti en tuple)
        # Dessin du centre du cluster en vert
        cv2.circle(frame_with_clusters, cluster_center, 10, (0, 255, 0), -1)
        # Ajout du texte "Cluster" près du centre
        cv2.putText(frame_with_clusters, 'Cluster', (cluster_center[0] - 30, cluster_center[1] - 30),
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)
        for fish_position in cluster_info["Fish Positions"]:  # Parcours de chaque position de poisson dans le cluster
            fish_position = tuple(map(int, fish_position))  # Convertir les coordonnées en tuple d'entiers
            # Dessin du cercle représentant la position du poisson en bleu
            cv2.circle(frame_with_clusters, fish_position, 5, (255, 0, 0), -1)
    return frame_with_clusters  # Renvoie l'image avec les clusters dessinés


def draw_clusters_on_video(video_path, clusters_info, output_path):
    """
    Dessine les clusters sur chaque trame d'une vidéo et enregistre la vidéo résultante.

    @input :
    - video_path (str): Chemin vers la vidéo d'entrée.
    - clusters_info (dict): Informations sur les clusters pour chaque trame de la vidéo.
    - output_path (str): Chemin de sortie pour enregistrer la vidéo avec les clusters dessinés.
    """
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        print("Error opening video file.")
        return

    frame_width = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    frame_height = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps = int(cap.get(cv2.CAP_PROP_FPS))

    # Define the codec and create VideoWriter object
    fourcc = cv2.VideoWriter_fourcc(*'mp4v')
    out = cv2.VideoWriter(output_path, fourcc, fps, (frame_width, frame_height))

    while True:
        ret, frame = cap.read()
        if not ret:
            break

        frame_number = int(cap.get(cv2.CAP_PROP_POS_FRAMES))
        if str(frame_number) in clusters_info:
            frame_clusters = clusters_info[str(frame_number)]["Clusters"]
            frame_with_clusters = draw_clusters_on_image(frame, frame_clusters)
            out.write(frame_with_clusters)
            cv2.imshow('Frame', frame_with_clusters)
        else:
            out.write(frame)
            cv2.imshow('Frame', frame)

        if cv2.waitKey(25) & 0xFF == ord('q'):
            break

    cap.release()
    out.release()
    cv2.destroyAllWindows()  # Fermer toutes les fenêtres OpenCV après la boucles


def plot_fish_positions(fish_centers, frame_number):
    """
    Trace les positions des poissons pour une trame donnée afin d'aider à déterminer les paramètres eps et min_samples de DBSCAN.

    @input :
    - fish_centers (dict): Dictionnaire contenant les coordonnées des poissons pour chaque trame.
    - frame_number (str): Numéro de la trame à visualiser.

    @output :
    - Visualisation des positions des poissons pour la trame spécifiée.
    """
    # Vérifier si le numéro de trame existe dans le dictionnaire
    if frame_number not in fish_centers:
        print("Le numéro de trame spécifié n'existe pas dans le dictionnaire.")
        return

    # Extraction des positions des poissons pour la trame spécifiée
    positions = [position[0] for position in fish_centers[frame_number].values()]

    # Conversion des positions en listes de coordonnées x et y séparées
    x_positions = [position[0] for position in positions]
    y_positions = [position[1] for position in positions]

    # Traçage des positions des poissons
    plt.figure(figsize=(8, 6))
    plt.scatter(x_positions, y_positions, color='blue')
    plt.title(f'Positions des poissons pour la trame {frame_number}')
    plt.xlabel('Position X')
    plt.ylabel('Position Y')
    plt.grid(True)
    plt.show()

def write_cluster_info_to_excel(data, excel_filename):
    """
    Écrit les informations sur les clusters dans un fichier Excel.

    @input :
    - data (dict): Données sur les clusters.
    - excel_filename (str): Nom du fichier Excel de sortie.
    """
    # Créer un nouveau classeur Excel
    wb = openpyxl.Workbook()

    # Sélectionner la première feuille de calcul
    sheet = wb.active

    # Ajouter des en-têtes
    sheet["A1"] = "Frame"
    sheet["B1"] = "Cluster"
    sheet["C1"] = "Nombre de poissons"

    # Parcourir les données et écrire dans le fichier Excel
    row = 2
    for frame, clusters_info in data.items():
        for cluster, info in clusters_info['Clusters'].items():
            num_poissons = len(info['Fish Positions'])
            sheet.cell(row=row, column=1, value=frame)
            sheet.cell(row=row, column=2, value=cluster)
            sheet.cell(row=row, column=3, value=num_poissons)
            row += 1
    
    # Enregistrer le classeur Excel
    wb.save(excel_filename)


def color_excel_cells(file_path):
    """
    Colorie les cellules dans un fichier Excel selon certaines conditions.

    @input :
    - file_path (str): Chemin vers le fichier Excel.
    """
    # Initialiser le style pour le remplissage jaune
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Charger le classeur Excel en mode écriture
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Charger les données du DataFrame
    df = pd.read_excel(file_path)
    
    # Parcourir les lignes du DataFrame
    for index, row in df.iterrows():
        # Vérifier si le nombre de poissons est égal à 1
        if row['Nombre de poissons'] == 1:
            # Colorier la cellule correspondante dans la colonne "Nombre de poissons" en jaune
            cell = ws.cell(row=index+2, column=3)  # Ajouter 2 à l'index pour correspondre à la ligne Excel (indexée à partir de 1)
            cell.fill = yellow_fill
    
    # Sauvegarder le classeur Excel avec les mises à jour
    wb.save(file_path)


def isolation_alert_threshold(file_path, alert_threshold):
    """
    Vérifie le pourcentage de clusters à un seul poisson et affiche une alerte si nécessaire.

    @input :
    - file_path (str): Chemin vers le fichier Excel.
    - alert_threshold (float): Seuil d'alerte pour le pourcentage de clusters à un seul poisson.
    """
    try:
        df = pd.read_excel(file_path)
        nombre_de_poissons_counts = df['Nombre de poissons'].value_counts()
        total_clusters = nombre_de_poissons_counts.sum()
        pourcentage_un_poisson = (nombre_de_poissons_counts.get(1, 0) / total_clusters) * 100
        if pourcentage_un_poisson > alert_threshold:
            root = tk.Tk()
            root.withdraw()
            messagebox.showwarning("Alerte", f"Le pourcentage de clusters à un seul poisson est de {pourcentage_un_poisson:.2f}%")
    except Exception as e:
        print(f"Une erreur s'est produite lors de la vérification du pourcentage de clusters : {str(e)}")


def calculate_fish_distribution(file_path):
    """
    Calcule la distribution des poissons dans un fichier Excel.

    @input :
    - file_path (str): Chemin vers le fichier Excel.

    @output :
    - result (dict): Distribution des poissons.
    """
    try:
        df = pd.read_excel(file_path)
        fish_count = df['Nombre de poissons'].value_counts()
        total_clusters = fish_count.sum()
        
        result = {}
        for num_fish, count in fish_count.items():
            percentage = (count / total_clusters) * 100
            result[num_fish] = {'count': count, 'percentage': percentage}
        
        return result
        
    except Exception as e:
        print(f"Une erreur s'est produite lors du calcul de la distribution des poissons : {str(e)}")


def plot_single_fish_clusters_pie_chart(file_path, save_path):
    """
    Affiche le diagramme circulaire de répartition des clusters par nombre de poissons.

    @input :
    - file_path (str): Chemin vers le fichier Excel.
    """
    try:
        df = pd.read_excel(file_path)
        fish_count = df['Nombre de poissons'].value_counts()
        total_clusters = fish_count.sum()
        
        labels = ['Multiples poissons', 'Poisson unique']
        sizes = [total_clusters - fish_count.get(1, 0), fish_count.get(1, 0)]
        colors = ['#ff9999','#66b3ff']
        explode = (0.1, 0) 
        plt.pie(sizes, explode=explode, labels=labels, colors=colors, autopct='%1.1f%%', shadow=True, startangle=140)
        plt.axis('equal')
        plt.title('Répartition des clusters par nombre de poissons')
        # Sauvegarder le diagramme
        plt.savefig(save_path)
        plt.show()
        
    except Exception as e:
        print(f"Une erreur s'est produite lors de la génération du diagramme circulaire : {str(e)}")


def plot_fish_distribution_pie_chart(fish_distribution, save_path):
    """
    Affiche le diagramme circulaire de répartition des clusters de poissons.

    @input :
    - fish_distribution (dict): Distribution des poissons.
    """
    try:
        labels = list(fish_distribution.keys())
        sizes = [fish_distribution[num_fish]['count'] for num_fish in labels]
        plt.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=140)
        plt.axis('equal')
        plt.title('Répartition des clusters de poissons')
        # Sauvegarder le diagramme
        plt.savefig(save_path)
        plt.show()
        
    except Exception as e:
        print(f"Une erreur s'est produite lors de la génération du diagramme circulaire : {str(e)}")

