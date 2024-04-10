from ultralytics import YOLO
import pandas as pd
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def read_excel_data(significant_fishs_path, clusters_info_path):
    try:
        # Lire les données des fichiers Excel
        significant_df = pd.read_excel(significant_fishs_path)
        clusters_df = pd.read_excel(clusters_info_path)
        
        
        return significant_df, clusters_df
        
    except Exception as e:
        print(f"Une erreur s'est produite lors de la lecture des fichiers Excel : {e}")
        return None, None
    
def find_frames_with_single_fish_and_significant_frames(significant_df, clusters_df):
    try:
        # Filtrer les frames où Nombre de poissons = 1 dans clusters_info.xlsx
        single_fish_frames = clusters_df[clusters_df["Nombre de poissons"] == 1]["Frame"].unique()

        # Filtrer les frames avec une valeur numérique dans la colonne Poisson isolé de 'significant_fishs.xlsx'
        significant_frames = significant_df[pd.to_numeric(significant_df["Poisson isolé"], errors='coerce').notnull()]["Frame"].unique()

        # Afficher les frames communs
        common_frames = set(single_fish_frames) & set(significant_frames)
        return single_fish_frames, significant_frames, common_frames

    except Exception as e:
        print(f"Une erreur s'est produite lors de la recherche des frames : {e}")
        return None, None, None
    
def excel_isolated_fish_creation(single_fish_frames, significant_frames, common_frames, output_file):

    data = {'Frame': list(range(0, max(single_fish_frames.max(), significant_frames.max()) + 1)),
                    'isolé': ['oui' if frame in common_frames else '' for frame in range(0, max(single_fish_frames.max(), significant_frames.max()) + 1)]}

    df = pd.DataFrame(data)

    df.to_excel(output_file, index=False)

    # Charger le fichier Excel
    wb = load_workbook(output_file)

    # Sélectionner la feuille de calcul active
    ws = wb.active

    # Colorier les cases
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws.max_row):
        for cell in row:
            if cell.value == 'oui':
                cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    wb.save(output_file)