
# Impotation des bibliothèques

import json
from sklearn.neighbors import NearestNeighbors
import numpy as np
from scipy import stats
import openpyxl
from openpyxl.styles import PatternFill
import tkinter as tk
from tkinter import messagebox



def create_dict_neighbors(fish_centers, output):
    """
    Crée un dictionnaire des voisins de chaque poisson à chaque frame.

    @input :
    - fish_centers (dict): Un dictionnaire des centres de chaque poisson avec leurs identifiants.
    - output (str): Le chemin de sortie pour sauvegarder le dictionnaire des voisins.

    @output :
    Le dictionnaire des voisins est sauvegardé dans un fichier.
    """

    # Déterminez le nombre de voisins à rechercher
    k = min(len(fish_centers) - 1, len(next(iter(fish_centers.values()))))  # Nombre de poissons - 1 ou le minimum

    dico_voisin = {}
    for frame, fishes in fish_centers.items():
        frame_data = {}
        if frame not in fish_centers:
            continue  # Ignorer les frames qui ne sont pas présentes dans le dictionnaire
        positions = np.array(list(fishes.values()))  # Récupérer les positions des poissons

        if len(positions.shape) > 1:
            positions = positions.reshape(-1, 2)  # Remodeler le tableau si nécessaire

        if k > 0 and len(positions) >= k:  # Vérifiez s'il y a suffisamment d'échantillons pour trouver des voisins
            nbrs = NearestNeighbors(n_neighbors=k, algorithm='ball_tree').fit(positions)
            distances, indices = nbrs.kneighbors(positions)
            for i, (fish_id, fish_position) in enumerate(fishes.items()):
                fish_neighbors = {}
                for j, neighbor_index in enumerate(indices[i]):
                    if neighbor_index != i:  # Exclure le poisson lui-même
                        neighbor_fish_id = list(fishes.keys())[neighbor_index]
                        neighbor_distance = distances[i][j]
                        fish_neighbors[f'voisin_id {neighbor_fish_id}'] = neighbor_distance
                frame_data[f'poisson {fish_id}'] = fish_neighbors
        dico_voisin[f'frame {frame}'] = frame_data

    # Export du dictionnaire dans un fichier JSON
    with open(output, 'w') as f:
        json.dump(dico_voisin, f)


def calculate_average_distances(dico_voisin,neighbors_json_path):
    """
    Calcule les distances moyennes entre chaque poisson et ses voisins.

    @input :
    - dico_voisin (dict): Le dictionnaire des voisins de chaque poisson à chaque frame.

    @output :
    - average_distances (dict): Un dictionnaire contenant les distances moyennes pour chaque poisson à chaque frame.
    """

    average_distances = {}
    for frame, frame_data in dico_voisin.items():
        frame_average_distances = {}
        for fish, neighbors in frame_data.items():
            fish_id = int(fish.split()[1])
            total_distance = sum(neighbors.values())
            average_distance = total_distance / len(neighbors)
            frame_average_distances[f'poisson {fish_id}'] = average_distance
        average_distances[frame] = frame_average_distances

    with open(neighbors_json_path, 'w') as f:
        json.dump(average_distances, f)
    return average_distances

def show_isolated_fish_percentage(isolated_frames, total_frames):
    """
    Affiche une alerte dans une fenêtre avec le pourcentage de la vidéo où il y a un poisson isolé.

    @input :
    - isolated_frames (int): Le nombre de frames où un poisson est isolé.
    - total_frames (int): Le nombre total de frames dans la vidéo.
    """
    percentage = (isolated_frames / total_frames) * 100
    message = f"Il y a un ou des poissons isolés pendant {percentage:.2f}% de la vidéo."
    root = tk.Tk()
    root.withdraw()
    messagebox.showwarning("Pourcentage de poissons isolés", message)


def calculate_isolated_fish_percentage(significant_fishs, total_frames):
    """
    Calcule le pourcentage de la vidéo où il y a un poisson isolé.

    @input :
    - significant_fishs (dict): Dictionnaire contenant les poissons significatifs par frame.
    - total_frames (int): Le nombre total de frames dans la vidéo.

    @output :
    - isolated_frames (int): Le nombre de frames où un poisson est isolé.
    """
    isolated_frames = len(significant_fishs)
    show_isolated_fish_percentage(isolated_frames, total_frames)


# Modifier votre fonction student_test pour qu'elle renvoie également le nombre total de frames
def student_test(json_file, alpha):
    """
    Effectue un test statistique sur les distances de chaque poisson à ses voisins.

    @input :
    - json_file (str): Le chemin du fichier JSON contenant les distances des voisins.
    - alpha (float): Le niveau de signification pour le test statistique.

    @output :
    - significant_fishs (dict): Dictionnaire contenant les poissons significatifs par frame.
    - total_frames (int): Le nombre total de frames dans la vidéo.
    """
    significant_fishs = {}
    total_frames = 0  # Initialisation du nombre total de frames

    with open(json_file, 'r') as f:
        data = json.load(f)

    for frame in data.keys():
        distances_values = list(data[frame].values())
        for fish in data[frame].keys():
            other_distances = []
            for elem in distances_values:
                if elem != data[frame][fish]:
                    other_distances.append(elem)

            mean = np.mean(other_distances)
            std_dev = np.std(other_distances)  # ecart-type
            degrees_of_freedom = len(other_distances)
            t_statistic = (data[frame][fish] - mean) / (std_dev / np.sqrt(degrees_of_freedom + 1))
            t_critical = stats.t.ppf(1 - alpha, degrees_of_freedom)

            if abs(t_statistic) > t_critical:
                dict_frame = frame.replace("frame", "").strip()
                dict_fish = fish.replace("poisson", "").strip()
                significant_fishs[dict_frame] = dict_fish

        total_frames += 1  # Incrémentation du nombre total de frames

    return significant_fishs, total_frames


def write_significant_fish_to_excel(significant_fishs, excel_filename):
    """
    Écrit les informations sur les poissons significatifs dans un fichier Excel.

    @input :
    - significant_fishs (dict): Dictionnaire contenant les poissons significatifs par frame.
    - excel_filename (str): Nom du fichier Excel de sortie.
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet["A1"] = "Frame"
    sheet["B1"] = "Poisson isolé"
    
    row = 2
    for frame, fish_id in significant_fishs.items():
        sheet.cell(row=row, column=1, value=frame)
        sheet.cell(row=row, column=2, value=fish_id)
        row += 1
    
    wb.save(excel_filename)


def color_excel_cells(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            if cell.value == 1:
                cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    wb.save(file_path)
